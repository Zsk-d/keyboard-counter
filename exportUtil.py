# Author:zsk-d

import shutil,xlwt,xlrd,time,os,sys
import openpyxl
from openpyxl.styles import PatternFill

COPY_EXCEL_PATH = 'data\\export_{}.xlsx'
TEMP_EXCEL_FILE_PATH = 'data\\export_temp.xlsx'
MAX_FILL_COLOR = 'ff0000'
MAX_KEY_PRESS_COUNT = 0

def getKeyDataXY(keyName):
    ''' 获取键盘按键对应excel模板位置 '''
    try:
        return {
        'Escape'  :(0,1),'F1'      :(2,1),'F2'       :(3,1),'F3'       :(4,1),'F4'    :(5,1),'F5'    :(6,1),'F6'        :(7,1),'F7'  :(8,1),'F8'       :(9,1),'F9'        :(10,1),'F10'  :(11,1),'F11'    :(12,1),'F12'       :(13,1),
        'Oem_3'   :(0,3),'1'       :(1,3),'2'        :(2,3),'3'        :(3,3),'4'     :(4,3),'5'     :(5,3),'6'         :(6,3),'7'   :(7,3),'8'        :(8,3),'9'         :(9,3),'0'     :(10,3),'Oem_Minus':(11,3),'Oem_Plus':(12,3),'Back' :(13,3),
        'Tab'     :(0,5),'Q'       :(1,5),'W'        :(2,5),'E'        :(3,5),'R'     :(4,5),'T'     :(5,5),'Y'         :(6,5),'U'   :(7,5),'I'        :(8,5),'O'         :(9,5),'P'     :(10,5),'Oem_4'  :(11,5),'Oem_6'     :(12,5),'Oem_5':(13,5),
        'Capital' :(0,7),'A'       :(2,7),'S'        :(3,7),'D'        :(4,7),'F'     :(5,7),'G'     :(6,7),'H'         :(7,7),'J'   :(8,7),'K'        :(9,7),'L'         :(10,7),'Oem_1':(11,7),'Oem_7'  :(12,7),'Return'    :(13,7),
        'Lshift'  :(0,9),'Z'       :(2,9),'X'        :(3,9),'C'        :(4,9),'V'     :(5,9),'B'     :(6,9),'N'         :(7,9),'M'   :(8,9),'Oem_Comma':(9,9),'Oem_Period':(10,9),'Oem_2':(11,9),'Rshift' :(12,9),
        'Lcontrol':(0,11),'Lwin'   :(1,11),'Lmenu'   :(2,11),'Space'   :(3,11),'Rmenu':(10,11),'Rwin':(11,11),'Rcontrol':(13,11),
        'Snapshot':(15,1),'Scroll' :(16,1),'Pause'   :(17,1),'Insert'  :(15,3),'Home' :(16,3),'Prior':(17,3),'Delete'   :(15,5),'End':(16,5),'Next'    :(17,5),
        'Up'      :(16,9),'Left'   :(15,11),'Down'   :(16,11),'Right'  :(17,11),
        'Numlock' :(19,3),'Divide' :(20,3),'Multiply':(21,3),'Subtract':(22,3),
        'Numpad7' :(19,5),'Numpad8':(20,5),'Numpad9' :(21,5),'Add'     :(22,5),
        'Numpad4' :(19,7),'Numpad5':(20,7),'Numpad6' :(21,7),
        'Numpad1' :(19,9),'Numpad2':(20,9),'Numpad3' :(21,9),'Numpad0' :(19,11),'Decimal':(21,11)
    }[keyName]
    except :
        return None
    
def getNowTimeString():
    return time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime(int(time.time()))) 

def setKeyPreeCount(keyRecordData):
    global MAX_KEY_PRESS_COUNT
    MAX_KEY_PRESS_COUNT = sorted(keyRecordData.items(),key=lambda x:x[1],reverse=True)[0][1]

def copyExportFile(filename:str):
    ''' 复制data文件夹下的键盘模板表格 '''
    if(not os.path.exists(filename)): return shutil.copyfile('data/export_temp.xlsx',filename)

def getFillColor(keyPressCount:int):
    if MAX_KEY_PRESS_COUNT == 0 or keyPressCount / MAX_KEY_PRESS_COUNT < 0.06 :return 'ffffff'
    else :return str(hex(int(MAX_FILL_COLOR,16) - (MAX_KEY_PRESS_COUNT - keyPressCount) * 100))[2:]

def exportImg(filename,sheet_name,screen_area):
    from win32com.client import DispatchEx
    import pythoncom
    from PIL import ImageGrab
    pic_name = filename[:-4] + 'png'
    pythoncom.CoInitialize()
    excel = DispatchEx("Excel.Application")
    excel.Visible = True
    excel.DisplayAlerts = False
    workbook = excel.Workbooks.Open(filename)
    worksheet = workbook.Sheets(sheet_name)
    worksheet.Range(screen_area).CopyPicture()
    worksheet.Paste()
    excel.Selection.ShapeRange.Name = pic_name
    worksheet.Shapes(pic_name).Copy()
    img = ImageGrab.grabclipboard()
    img.save(pic_name)
    workbook.Close(SaveChanges=0)
    excel.Quit()
    pythoncom.CoUninitialize()

def exportKeyRecordExcel(keyRecordData:tuple):
    ''' 导出键盘统计数据 '''
    # 获取按键次数极值
    setKeyPreeCount(keyRecordData)
    # 获取文件名
    filename = COPY_EXCEL_PATH.format(getNowTimeString())
    # 复制excel文件 
    workerbook = openpyxl.load_workbook(TEMP_EXCEL_FILE_PATH)
    worksheet = workerbook.get_sheet_by_name('export')
    for key in keyRecordData.keys():
        if keyRecordData[key] == 'null': continue
        keyMapExcelXY = getKeyDataXY(key)
        if keyMapExcelXY:
            row = keyMapExcelXY[1] + 1
            col = keyMapExcelXY[0] + 1
            fill = PatternFill(start_color = getFillColor(keyRecordData[key]), fill_type = 'solid')
            worksheet.cell(row,col,str(keyRecordData[key])).fill = fill
    workerbook.save(filename)
    exportImg(os.path.dirname(sys.argv[0]) + '/' + filename,'export','A1:W12')
    # 打开文件目录并选定
    openFilePath = 'explorer /e,/select,{}'.format(os.path.dirname(sys.argv[0]) + '\\' + filename)
    os.popen(openFilePath)